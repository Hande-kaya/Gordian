"""
Excel Extraction Service - Read Excel files and extract data via LLM.

Excel files don't need OCR — cell data is read directly with openpyxl/xlrd,
then the text is sent through the same LLM extraction pipeline as PDF/images.
Supports both invoice/income and bank-statement doc types.
"""

import io
import json
import logging
import os
import re
from typing import Dict, Any, Optional, List

logger = logging.getLogger(__name__)

MAX_ROWS = 500
MAX_COLS = 30


def _read_xlsx(file_bytes: bytes) -> Optional[str]:
    """Read .xlsx file and return pipe-delimited text."""
    import openpyxl

    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), read_only=True, data_only=True
    )
    ws = wb.active
    if ws is None:
        wb.close()
        return None

    lines = []
    for row_idx, row in enumerate(ws.iter_rows(max_col=MAX_COLS, values_only=True)):
        if row_idx >= MAX_ROWS:
            break
        values = [str(cell) if cell is not None else '' for cell in row]
        # Trim trailing empty cells
        while values and not values[-1].strip():
            values.pop()
        if not values:
            continue
        lines.append(' | '.join(values))

    wb.close()
    return '\n'.join(lines)


def _read_xls(file_bytes: bytes) -> Optional[str]:
    """Read legacy .xls file and return pipe-delimited text."""
    import xlrd

    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)

    lines = []
    max_row = min(ws.nrows, MAX_ROWS)
    max_col = min(ws.ncols, MAX_COLS)

    for row_idx in range(max_row):
        values = []
        for col_idx in range(max_col):
            cell = ws.cell(row_idx, col_idx)
            values.append(str(cell.value) if cell.value is not None else '')
        # Trim trailing empty cells
        while values and not values[-1].strip():
            values.pop()
        if not values:
            continue
        lines.append(' | '.join(values))

    return '\n'.join(lines)


def _read_excel_text(file_bytes: bytes, filename: str) -> Optional[str]:
    """Read Excel file and return pipe-delimited text."""
    ext = os.path.splitext(filename)[1].lower()
    if ext == '.xls':
        return _read_xls(file_bytes)
    return _read_xlsx(file_bytes)


def extract_from_bytes(
    file_bytes: bytes, filename: str, doc_type: str = 'invoice'
) -> Dict[str, Any]:
    """Extract data from Excel file bytes.

    Args:
        file_bytes: Raw file content
        filename: Original filename (used to detect .xlsx vs .xls)
        doc_type: Document type ('invoice', 'income', 'bank-statement', etc.)

    Returns:
        dict with 'extracted_text' and 'extracted_data'
    """
    text = _read_excel_text(file_bytes, filename)

    if not text or not text.strip():
        logger.warning(f"Empty Excel content for {filename}")
        return {'extracted_text': '', 'extracted_data': {}}

    logger.info(
        f"Excel read for {filename}: {len(text)} chars, "
        f"{text.count(chr(10)) + 1} rows"
    )

    if doc_type == 'bank-statement':
        extracted_data = _extract_bank_statement(text)
    else:
        from services.llm_extraction_service import get_llm_extraction_service
        llm_service = get_llm_extraction_service()
        extracted_data = llm_service._llm_extract_entities(text)

    return {
        'extracted_text': text,
        'extracted_data': extracted_data,
    }


def _extract_bank_statement(text: str) -> Dict[str, Any]:
    """Extract bank statement data from Excel text via LLM.

    Excel bank statement data is clean and structured (no OCR noise),
    so LLM reliably extracts both header fields and transactions.
    """
    from openai import OpenAI
    from services.bank_statement_utils import (
        safe_float, detect_currency_from_text,
        normalize_currency as _normalize_currency,
    )
    from services.bank_statement_extractor import (
        _compute_totals, _convert_header, empty_bank_statement_data,
    )

    prompt = f"""Extract ALL information from this bank/credit card statement.

TEXT:
{text[:12000]}

RESPOND WITH JSON ONLY:
{{
    "bank_name": "full bank name",
    "account_holder": "account holder name or null",
    "account_number": "IBAN, account number, or customer ID",
    "card_number": "masked card number or null",
    "statement_date": "YYYY-MM-DD (statement issue date)",
    "statement_period_start": "YYYY-MM-DD",
    "statement_period_end": "YYYY-MM-DD",
    "currency": "3-letter ISO 4217 code (EUR, USD, TRY, etc). Convert: €=EUR, $=USD, TL=TRY. Return null if uncertain.",
    "opening_balance": number or null,
    "closing_balance": number or null,
    "transactions": [
        {{
            "date": "YYYY-MM-DD",
            "description": "transaction description",
            "type": "debit or credit",
            "amount": positive_number,
            "balance": number_or_null
        }}
    ]
}}

RULES:
- All amounts as positive numbers (use type field for debit/credit)
- Negative amounts or payments out = debit, positive amounts or payments in = credit
- Dates in YYYY-MM-DD format
- Include ALL transactions found
- Return null for fields not found"""

    try:
        client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=4000,
            response_format={"type": "json_object"}
        )
        result = json.loads(response.choices[0].message.content)
        logger.info(f"Bank stmt LLM extracted: {len(result.get('transactions', []))} txs")

        # Build header
        header = _convert_header(result)

        # Process transactions
        transactions = []
        for tx in result.get('transactions', []):
            if not isinstance(tx, dict):
                continue
            amount = safe_float(tx.get('amount'))
            if amount is None:
                continue
            transactions.append({
                'date': tx.get('date'),
                'description': tx.get('description', ''),
                'type': tx.get('type', 'debit'),
                'amount': amount,
                'balance': safe_float(tx.get('balance')),
            })

        header['transactions'] = transactions
        _compute_totals(header, transactions)

        # Currency fallback
        if not header.get('currency'):
            detected = detect_currency_from_text(text)
            if detected:
                header['currency'] = detected

        # Normalize transaction amounts
        try:
            from services.exchange_rate_service import normalize_transaction
            stmt_currency = header.get('currency')
            for tx in transactions:
                normalize_transaction(tx, stmt_currency)
        except Exception as e:
            logger.warning(f"Excel bank tx normalization failed: {e}")

        return header

    except Exception as e:
        logger.error(f"Excel bank statement extraction error: {e}")
        return empty_bank_statement_data()
